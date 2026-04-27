from __future__ import annotations

import csv
import json
import re
import zipfile
from html.parser import HTMLParser
from pathlib import Path
from typing import Any


class _HTMLText(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows: list[str] = []
    def handle_data(self, data: str) -> None:
        text = data.strip()
        if text:
            self.rows.append(text)


class DocumentToolbox:
    toolbox_name = "docs"
    tags = ("builtin", "document", "extract")

    def __init__(self, workspace_root: Path | None = None):
        self.workspace_root = workspace_root.resolve() if workspace_root else None
        self.runtime = None

    def bind_runtime(self, runtime, tool_lookup=None) -> None:  # noqa: ANN001
        self.runtime = runtime

    def spawn(self, workspace_root: Path) -> "DocumentToolbox":
        return DocumentToolbox(workspace_root=workspace_root)

    def executors(self):
        return {
            "docs.extract_text": self._exec_extract_text,
            "docs.metadata": self._exec_metadata,
            "docs.table_preview": self._exec_table_preview,
        }

    def _root(self) -> Path:
        if self.workspace_root is None:
            raise ValueError("DocumentToolbox workspace not bound yet.")
        return self.workspace_root.resolve()

    def _safe_path(self, raw: str) -> Path:
        path = (self._root() / str(raw or "")).resolve()
        if not path.is_relative_to(self._root()):
            raise ValueError(f"Path escapes workspace: {raw}")
        return path

    def _rel(self, path: Path) -> str:
        return path.resolve().relative_to(self._root()).as_posix()

    def _exec_metadata(self, args: dict[str, Any]):
        path = self._safe_path(args["path"])
        stat = path.stat()
        return json.dumps({"path": self._rel(path), "suffix": path.suffix.lower(), "size": stat.st_size, "modified_at": int(stat.st_mtime)}, ensure_ascii=False, indent=2)

    def _exec_extract_text(self, args: dict[str, Any]):
        path = self._safe_path(args["path"])
        max_chars = max(1, min(int(args.get("max_chars", 60000) or 60000), 300000))
        text = self._extract(path)
        return text[:max_chars]

    def _extract(self, path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix in {".txt", ".md", ".markdown", ".json", ".yaml", ".yml", ".toml", ".csv", ".py", ".js", ".ts", ".html", ".xml"}:
            text = path.read_text(encoding="utf-8", errors="replace")
            if suffix in {".html", ".xml"}:
                parser = _HTMLText(); parser.feed(text); return "\n".join(parser.rows)
            return text
        if suffix == ".docx":
            return self._extract_docx(path)
        if suffix == ".xlsx":
            return self._extract_xlsx(path)
        if suffix == ".pdf":
            return self._extract_pdf(path)
        raise ValueError(f"Unsupported document type: {suffix}")

    def _extract_docx(self, path: Path) -> str:
        with zipfile.ZipFile(path) as zf:
            xml = zf.read("word/document.xml").decode("utf-8", errors="replace")
        xml = re.sub(r"</w:p>", "\n", xml)
        return re.sub(r"<[^>]+>", "", xml).replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")

    def _extract_xlsx(self, path: Path) -> str:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError("XLSX extraction requires openpyxl") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        rows = []
        for ws in wb.worksheets[:5]:
            rows.append(f"# Sheet: {ws.title}")
            for row in ws.iter_rows(max_row=80, values_only=True):
                rows.append("\t".join("" if cell is None else str(cell) for cell in row))
        return "\n".join(rows)

    def _extract_pdf(self, path: Path) -> str:
        try:
            from pypdf import PdfReader  # type: ignore
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError("PDF extraction requires pypdf in offline environment") from exc
        reader = PdfReader(str(path))
        return "\n\n".join(page.extract_text() or "" for page in reader.pages[:30])

    def _exec_table_preview(self, args: dict[str, Any]):
        path = self._safe_path(args["path"])
        suffix = path.suffix.lower()
        limit = max(1, min(int(args.get("limit", 20) or 20), 100))
        if suffix == ".csv":
            with path.open("r", encoding="utf-8", errors="replace", newline="") as fh:
                rows = [row for _, row in zip(range(limit), csv.reader(fh))]
            return json.dumps({"path": self._rel(path), "rows": rows}, ensure_ascii=False, indent=2)
        if suffix == ".xlsx":
            text = self._extract_xlsx(path)
            return "\n".join(text.splitlines()[:limit])
        raise ValueError("table_preview supports CSV/XLSX only")
